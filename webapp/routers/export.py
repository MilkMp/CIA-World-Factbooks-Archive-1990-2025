"""Data export routes — CSV and Excel downloads."""

import csv
import io
from pathlib import Path

from fastapi import APIRouter, Request
from fastapi.responses import StreamingResponse
from fastapi.templating import Jinja2Templates

from webapp.database import sql, sql_one
from webapp.parsers import (
    extract_number, extract_pct_gdp, extract_pct,
    extract_gdp_percap, extract_dollar_billions,
    parse_life_exp, extract_rate, extract_growth_rate,
    extract_per_100, extract_area, extract_total_subs,
)

router = APIRouter()
templates = Jinja2Templates(directory=Path(__file__).resolve().parent.parent / "templates")

# ── Typed value extraction for exports ────────────────────
# Maps CanonicalName -> (parser_func, excel_number_format)
FIELD_PARSERS = {
    'Population': (extract_number, '#,##0'),
    'Area': (extract_area, '#,##0'),
    'Real GDP (purchasing power parity)': (extract_dollar_billions, '$#,##0.0'),
    'Real GDP per capita': (extract_gdp_percap, '$#,##0'),
    'GDP (purchasing power parity)': (extract_dollar_billions, '$#,##0.0'),
    'GDP - per capita': (extract_gdp_percap, '$#,##0'),
    'GDP - per capita (PPP)': (extract_gdp_percap, '$#,##0'),
    'Life expectancy at birth': (parse_life_exp, '0.0'),
    'Military expenditures': (extract_pct_gdp, '0.0%'),
    'Population growth rate': (extract_growth_rate, '0.00%'),
    'Birth rate': (extract_rate, '0.0'),
    'Death rate': (extract_rate, '0.0'),
    'Infant mortality rate': (extract_rate, '0.0'),
    'Internet users': (extract_pct, '0.0%'),
    'Unemployment rate': (extract_pct, '0.0%'),
    'Inflation rate (consumer prices)': (extract_pct, '0.0%'),
    'Telephones - mobile cellular': (extract_total_subs, '#,##0'),
    'Telephones - fixed lines': (extract_total_subs, '#,##0'),
    'Broadband - fixed subscriptions': (extract_total_subs, '#,##0'),
}

# Percent fields need value / 100 for Excel's native % format
_PCT_FORMATS = {'0.0%', '0.00%'}


def _parse_typed_value(canonical_name: str, content: str):
    """Try to extract a typed numeric value from a field's content.

    Returns (value, excel_format) or (None, None).
    """
    if not canonical_name or not content:
        return None, None
    entry = FIELD_PARSERS.get(canonical_name)
    if not entry:
        return None, None
    parser, fmt = entry
    val = parser(content)
    if val is None:
        return None, None
    # Excel native % format expects 0.05 for 5%, so divide by 100
    if fmt in _PCT_FORMATS:
        val = val / 100.0
    return val, fmt

GITHUB_REPO = "https://github.com/MilkMp/CIA-World-Factbooks-Archive-1990-2025"


def _safe_filename(name: str) -> str:
    return name.replace(" ", "_").replace(",", "").replace("'", "")


def _lookup_country(year: int, code: str):
    return sql_one("""
        SELECT c.CountryID, c.Year, mc.CanonicalName, mc.ISOAlpha2,
               mc.MasterCountryID
        FROM Countries c
        JOIN MasterCountries mc ON c.MasterCountryID = mc.MasterCountryID
        WHERE c.Year = ? AND (mc.CanonicalCode = ? OR mc.ISOAlpha2 = ?)
    """, [year, code.upper(), code.upper()])


def _lookup_master(code: str):
    return sql_one("""
        SELECT MasterCountryID, CanonicalName, ISOAlpha2
        FROM MasterCountries
        WHERE CanonicalCode = ? OR ISOAlpha2 = ?
    """, [code.upper(), code.upper()])


def _get_fields(country_id: int):
    return sql("""
        SELECT cc.CategoryTitle, cf.FieldName, cf.Content, fm.CanonicalName
        FROM CountryFields cf
        JOIN CountryCategories cc ON cf.CategoryID = cc.CategoryID
        LEFT JOIN FieldNameMappings fm ON cf.FieldName = fm.OriginalName
        WHERE cf.CountryID = ?
        ORDER BY cc.CategoryTitle, cf.FieldName
    """, [country_id])


def _get_bulk_fields(master_id: int):
    return sql("""
        SELECT c.Year, cc.CategoryTitle, cf.FieldName, cf.Content, fm.CanonicalName
        FROM CountryFields cf
        JOIN Countries c ON cf.CountryID = c.CountryID
        JOIN CountryCategories cc ON cf.CategoryID = cc.CategoryID
        LEFT JOIN FieldNameMappings fm ON cf.FieldName = fm.OriginalName
        WHERE c.MasterCountryID = ?
        ORDER BY c.Year, cc.CategoryTitle, cf.FieldName
    """, [master_id])


# ── HTML page ──────────────────────────────────────────────

@router.get("/export")
async def export_page(request: Request):
    countries = sql("""
        SELECT mc.MasterCountryID, mc.CanonicalName, mc.ISOAlpha2,
               mc.CanonicalCode, mc.EntityType,
               COUNT(DISTINCT c.Year) AS YearCount
        FROM MasterCountries mc
        JOIN Countries c ON mc.MasterCountryID = c.MasterCountryID
        GROUP BY mc.MasterCountryID, mc.CanonicalName, mc.ISOAlpha2,
                 mc.CanonicalCode, mc.EntityType
        ORDER BY mc.CanonicalName
    """)
    return templates.TemplateResponse("archive/export.html", {
        "request": request,
        "countries": countries,
        "github_repo": GITHUB_REPO,
    })


# ── Printable report ──────────────────────────────────────

@router.get("/export/print")
async def export_print(request: Request, codes: str = "", start: int = 1990, end: int = 2025):
    """Render a formatted, print-ready page for one or more countries across a year range."""
    code_list = [c.strip().upper() for c in codes.split(",") if c.strip()]
    if not code_list:
        return templates.TemplateResponse("archive/print.html", {
            "request": request, "entries": [], "error": "No country selected.",
        })

    start = max(1990, min(start, 2025))
    end = max(start, min(end, 2025))

    entries = []
    for code in code_list:
        master = _lookup_master(code)
        if not master:
            continue

        # Get all country rows in the year range
        years = sql("""
            SELECT CountryID, Year FROM Countries
            WHERE MasterCountryID = ? AND Year BETWEEN ? AND ?
            ORDER BY Year
        """, [master['MasterCountryID'], start, end])

        for yr in years:
            fields = _get_fields(yr['CountryID'])
            categories = {}
            for f in fields:
                cat = f['CategoryTitle'] or 'Uncategorized'
                categories.setdefault(cat, []).append(f)

            entries.append({
                'name': master['CanonicalName'],
                'iso2': master['ISOAlpha2'],
                'year': yr['Year'],
                'categories': categories,
            })

    return templates.TemplateResponse("archive/print.html", {
        "request": request,
        "entries": entries,
        "error": None,
    })


# ── Bulk exports (all years) — must be before single-year routes ──

@router.get("/export/bulk/{code}/csv")
async def export_bulk_csv(code: str):
    master = _lookup_master(code)
    if not master:
        return StreamingResponse(io.BytesIO(b"Country not found"),
                                 media_type="text/plain", status_code=404)

    fields = _get_bulk_fields(master['MasterCountryID'])

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Year", "Category", "Field Name", "Content", "Value"])
    for f in fields:
        canon = f['CanonicalName'] or f['FieldName']
        val, _ = _parse_typed_value(canon, f['Content'])
        writer.writerow([
            f['Year'],
            f['CategoryTitle'] or 'Uncategorized',
            canon,
            f['Content'],
            val if val is not None else '',
        ])

    filename = f"{_safe_filename(master['CanonicalName'])}_1990-2025_all_years.csv"
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/bulk/{code}/xlsx")
async def export_bulk_xlsx(code: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, numbers

    master = _lookup_master(code)
    if not master:
        return StreamingResponse(io.BytesIO(b"Country not found"),
                                 media_type="text/plain", status_code=404)

    fields = _get_bulk_fields(master['MasterCountryID'])

    wb = Workbook()
    ws = wb.active
    ws.title = master['CanonicalName'][:31]  # Excel sheet name max 31 chars

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1C2127", end_color="1C2127", fill_type="solid")
    for col, h in enumerate(["Year", "Category", "Field Name", "Content", "Value"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill

    for i, f in enumerate(fields, 2):
        canon = f['CanonicalName'] or f['FieldName']
        ws.cell(row=i, column=1, value=f['Year'])
        ws.cell(row=i, column=2, value=f['CategoryTitle'] or 'Uncategorized')
        ws.cell(row=i, column=3, value=canon)
        ws.cell(row=i, column=4, value=f['Content'])
        val, fmt = _parse_typed_value(canon, f['Content'])
        if val is not None:
            cell = ws.cell(row=i, column=5, value=val)
            cell.number_format = fmt

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{_safe_filename(master['CanonicalName'])}_1990-2025_all_years.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Single-year exports ───────────────────────────────────

@router.get("/export/{code}/{year}/csv")
async def export_csv(year: int, code: str):
    country = _lookup_country(year, code)
    if not country:
        return StreamingResponse(io.BytesIO(b"Country not found"),
                                 media_type="text/plain", status_code=404)

    fields = _get_fields(country['CountryID'])

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Category", "Field Name", "Content", "Value"])
    for f in fields:
        canon = f['CanonicalName'] or f['FieldName']
        val, _ = _parse_typed_value(canon, f['Content'])
        writer.writerow([
            f['CategoryTitle'] or 'Uncategorized',
            canon,
            f['Content'],
            val if val is not None else '',
        ])

    filename = f"{_safe_filename(country['CanonicalName'])}_{year}.csv"
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/{code}/{year}/xlsx")
async def export_xlsx(year: int, code: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, numbers

    country = _lookup_country(year, code)
    if not country:
        return StreamingResponse(io.BytesIO(b"Country not found"),
                                 media_type="text/plain", status_code=404)

    fields = _get_fields(country['CountryID'])

    wb = Workbook()
    ws = wb.active
    ws.title = f"{country['CanonicalName']} {year}"

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1C2127", end_color="1C2127", fill_type="solid")
    for col, h in enumerate(["Category", "Field Name", "Content", "Value"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill

    for i, f in enumerate(fields, 2):
        canon = f['CanonicalName'] or f['FieldName']
        ws.cell(row=i, column=1, value=f['CategoryTitle'] or 'Uncategorized')
        ws.cell(row=i, column=2, value=canon)
        ws.cell(row=i, column=3, value=f['Content'])
        val, fmt = _parse_typed_value(canon, f['Content'])
        if val is not None:
            cell = ws.cell(row=i, column=4, value=val)
            cell.number_format = fmt

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{_safe_filename(country['CanonicalName'])}_{year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
